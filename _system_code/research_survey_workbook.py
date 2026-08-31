from __future__ import annotations

import argparse
import getpass
import json
import os
import platform
import socket
import uuid
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

try:
    import psutil
except ImportError:  # pragma: no cover - optional student environment dependency
    psutil = None


DEFAULT_BASE_URL = "https://met-employability-services.azurewebsites.net"
DEFAULT_PATH_PREFIX = "/met-career-match"
NOTEBOOK_VERSION = "research-survey-workbook-v1"
ALLOWED_CHAT_HOSTS = {
    "claude.ai",
    "chatgpt.com",
    "chat.openai.com",
    "gemini.google.com",
    "aistudio.google.com",
    "copilot.microsoft.com",
    "perplexity.ai",
}
PLACEHOLDER_TEXT = {"", "na", "n/a", "none", "null", "todo", "test", "asdf", "qwerty"}


def load_dotenv_file(path=".env"):
    env_path = Path(path)
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def api_key_from_env():
    api_key = os.environ.get("EMPLOYABILITY_API_KEY", "").strip()
    if api_key:
        return api_key
    return getpass.getpass("Enter your Student API key: ")


def base_url_from_env():
    return os.environ.get("EMPLOYABILITY_API_BASE_URL", DEFAULT_BASE_URL).rstrip("/")


def path_prefix_from_env(base_url):
    prefix = os.environ.get("EMPLOYABILITY_API_PATH_PREFIX", DEFAULT_PATH_PREFIX).strip().strip("/").strip()
    if prefix and base_url.rstrip("/").endswith("/" + prefix):
        return ""
    return prefix


def build_api_url(base_url, prefix, path):
    clean_path = path.strip("/")
    clean_prefix = str(prefix or "").strip("/")
    if clean_prefix:
        return f"{base_url}/{clean_prefix}/{clean_path}/"
    return f"{base_url}/{clean_path}/"


def _with_port(parsed_url, port):
    netloc = parsed_url.hostname or ""
    if ":" in netloc and not netloc.startswith("["):
        netloc = f"[{netloc}]"
    if parsed_url.username:
        auth = parsed_url.username
        if parsed_url.password:
            auth = f"{auth}:{parsed_url.password}"
        netloc = f"{auth}@{netloc}"
    netloc = f"{netloc}:{port}"
    return urlunparse(
        (
            parsed_url.scheme,
            netloc,
            parsed_url.path.rstrip("/"),
            "",
            "",
            "",
        )
    )


def base_url_candidates(base_url):
    candidates = [base_url]
    parsed = urlparse(base_url)
    if parsed.scheme == "http" and parsed.hostname in {"127.0.0.1", "localhost"} and parsed.port == 9000:
        local_django_url = _with_port(parsed, 8000)
        if local_django_url not in candidates:
            candidates.append(local_django_url)
    return candidates


def survey_url_candidates(base_url, prefix, assignment_code):
    candidates = []
    for candidate_base_url in base_url_candidates(base_url):
        for candidate_prefix in [prefix, ""]:
            candidate = build_api_url(candidate_base_url, candidate_prefix, f"api/v1/research-surveys/{assignment_code}")
            if candidate not in candidates:
                candidates.append(candidate)
    return candidates


def survey_submit_url_candidates(base_url, prefix, assignment_code):
    candidates = []
    for candidate_base_url in base_url_candidates(base_url):
        candidate = build_api_url(candidate_base_url, prefix, f"api/v1/research-surveys/{assignment_code}/submit")
        if candidate not in candidates:
            candidates.append(candidate)
    return candidates


def survey_failure_message(attempts, action="reach"):
    guidance = [
        f"Survey request failed. The workbook could not {action} the research survey API.",
        "Check EMPLOYABILITY_API_BASE_URL in your .env file. For the course API, use "
        f"{DEFAULT_BASE_URL} with EMPLOYABILITY_API_PATH_PREFIX=/met-career-match.",
        "For local Django runserver, use http://127.0.0.1:8000/met-career-match with "
        "EMPLOYABILITY_API_PATH_PREFIX left empty.",
    ]
    joined_errors = " ".join(str(attempt.get("error", "")) for attempt in attempts)
    joined_urls = " ".join(str(attempt.get("url", "")) for attempt in attempts)
    if "BadStatusLine" in joined_errors and "127.0.0.1:9000" in joined_urls:
        guidance.append(
            "The 127.0.0.1:9000 port answered with non-HTTP data, so another local service is probably using that "
            "port or the server protocol is not plain HTTP."
        )
    return "\n".join(guidance) + f"\nAttempts: {json.dumps(attempts, indent=2)}"


def fetch_survey(assignment_code):
    load_dotenv_file()
    base_url = base_url_from_env()
    prefix = path_prefix_from_env(base_url)
    headers = {"X-API-Key": api_key_from_env(), "Accept": "application/json"}
    attempts = []
    for url in survey_url_candidates(base_url, prefix, assignment_code):
        try:
            response = requests.get(url, headers=headers, timeout=30)
        except requests.exceptions.RequestException as exc:
            attempts.append({"url": url, "error": str(exc)})
            continue
        preview = response.text[:240]
        attempts.append({"url": url, "status_code": response.status_code, "preview": preview})
        if response.status_code == 200:
            try:
                return response.json(), url
            except ValueError:
                attempts[-1]["error"] = f"Response was not valid JSON: {preview}"
                continue
    raise RuntimeError(survey_failure_message(attempts))


def detect_aws_instance(timeout=0.3):
    metadata_base = "http://169.254.169.254/latest"
    try:
        token_response = requests.put(
            f"{metadata_base}/api/token",
            headers={"X-aws-ec2-metadata-token-ttl-seconds": "60"},
            timeout=timeout,
        )
        if token_response.status_code != 200:
            return {"is_aws_ec2": False}
        headers = {"X-aws-ec2-metadata-token": token_response.text}
        fields = {
            "instance_id": "meta-data/instance-id",
            "instance_type": "meta-data/instance-type",
            "availability_zone": "meta-data/placement/availability-zone",
            "public_hostname": "meta-data/public-hostname",
            "local_ipv4": "meta-data/local-ipv4",
        }
        metadata = {"is_aws_ec2": True}
        for key, path in fields.items():
            try:
                response = requests.get(f"{metadata_base}/{path}", headers=headers, timeout=timeout)
                metadata[key] = response.text if response.status_code == 200 else ""
            except requests.exceptions.RequestException:
                metadata[key] = ""
        return metadata
    except requests.exceptions.RequestException:
        return {"is_aws_ec2": False}


def runtime_report():
    report = {
        "captured_at": datetime.now(timezone.utc).isoformat(),
        "hostname": socket.gethostname(),
        "operating_system": f"{platform.system()} {platform.release()}",
        "python_version": platform.python_version(),
        "machine": platform.machine(),
        "processor": platform.processor(),
        "aws": detect_aws_instance(),
    }
    if psutil:
        mem = psutil.virtual_memory()
        report["memory_total_mb"] = round(mem.total / (1024**2), 2)
        report["memory_available_mb"] = round(mem.available / (1024**2), 2)
    else:
        report["memory_note"] = "Install psutil for memory reporting."
    return report


def survey_items_in_order(survey_payload):
    return [item for section in survey_payload["sections"] for item in section["items"]]


def item_instructions(item):
    if item["response_type"] == "likert_5":
        return f"Enter an integer from {item.get('scale_min') or 1} to {item.get('scale_max') or 5}."
    if item["response_type"] == "yes_no":
        return "Enter TRUE or FALSE."
    if item["response_type"] == "text":
        return "Enter a short written response."
    return f"Unsupported response type: {item['response_type']}"


def style_header(row):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_validation(ws, response_type, cell_ref):
    if response_type == "likert_5":
        validation = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=False)
    elif response_type == "yes_no":
        validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    else:
        return
    ws.add_data_validation(validation)
    validation.add(ws[cell_ref])


def create_workbook(survey_payload, workbook_path, *, force=False):
    workbook_path = Path(workbook_path)
    if workbook_path.exists() and not force:
        return False
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    runtime_context = runtime_report()

    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    instructions.append(["Research response workbook"])
    instructions.append(["Fill only the Value cells. Do not change item_code values. Do not include your API key."])
    instructions.append(["Likert scale", "1 = Strongly disagree; 2 = Disagree; 3 = Neither agree nor disagree; 4 = Agree; 5 = Strongly agree"])
    instructions.append(["After editing", "Save this Excel file, then run the notebook/script validation step."])
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 110
    instructions["A1"].font = Font(bold=True, size=14)

    survey_sheet = wb.create_sheet("Survey")
    survey_sheet.append(["item_code", "section", "construct", "source_label", "response_type", "prompt", "instructions", "value"])
    style_header(survey_sheet[1])
    for item in survey_items_in_order(survey_payload):
        survey_sheet.append(
            [
                item["item_code"],
                item["section"],
                item.get("construct", ""),
                item.get("source_label", ""),
                item["response_type"],
                item["prompt"],
                item_instructions(item),
                None,
            ]
        )
        add_validation(survey_sheet, item["response_type"], f"H{survey_sheet.max_row}")
    survey_sheet.freeze_panes = "A2"
    widths = {"A": 24, "B": 18, "C": 30, "D": 32, "E": 16, "F": 76, "G": 34, "H": 24}
    for col, width in widths.items():
        survey_sheet.column_dimensions[col].width = width

    assignment_sheet = wb.create_sheet("Assignment")
    assignment_sheet.append(["field", "value", "notes"])
    style_header(assignment_sheet[1])
    assignment_rows = [
        ("data_pull_summary.raw_jobs_pulled", None, "Number, if applicable"),
        ("data_pull_summary.bronze_records_written", None, "Number, if applicable"),
        ("data_pull_summary.silver_records_written", None, "Number, if applicable"),
        ("data_pull_summary.duplicate_jobs_removed", None, "Number, if applicable"),
        ("data_pull_summary.malformed_rows_excluded", None, "Number, if applicable"),
        ("data_pull_summary.aws_hours_used", None, "Estimated EC2/runtime hours used for this assignment"),
        ("data_pull_summary.estimated_aws_spend_usd", None, "Estimated AWS spend in USD for this assignment"),
        ("domain_topic.target_role", "", "Target role/domain"),
        ("domain_topic.target_industry", "", "Target industry"),
        ("domain_topic.target_geography", "", "Target geography"),
        ("domain_topic.skill_cluster", "", "Target skill cluster"),
        ("selected_resources", "", "Comma-separated resources, e.g., BLS, O*NET, FRED"),
        ("assignment_answers.thought_process", "", "At least 5 words"),
        ("assignment_answers.assignment_usefulness", "", "At least 3 words"),
        ("assignment_answers.aws_cost_notes", "", "Optional note on how you estimated runtime/cost"),
        ("output_file_manifest", "", "Comma-separated output file paths"),
    ]
    for row in assignment_rows:
        assignment_sheet.append(list(row))
    assignment_sheet.column_dimensions["A"].width = 42
    assignment_sheet.column_dimensions["B"].width = 62
    assignment_sheet.column_dimensions["C"].width = 52

    ai_sheet = wb.create_sheet("AI Chat")
    ai_sheet.append(["field", "value", "notes"])
    style_header(ai_sheet[1])
    for row in [
        ("used_genai", "", "TRUE if you used GenAI for this assignment; otherwise FALSE"),
        ("provider", "", "claude, chatgpt, gemini, copilot, perplexity, or other"),
        ("chat_url", "", "Required only when used_genai is TRUE. Must be HTTPS shared-chat URL."),
        ("student_confirmed_reviewed", "", "TRUE after checking the shared chat for sensitive information."),
    ]:
        ai_sheet.append(list(row))
    for cell_ref in ["B2", "B5"]:
        validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ai_sheet.add_data_validation(validation)
        validation.add(ai_sheet[cell_ref])
    ai_sheet.column_dimensions["A"].width = 34
    ai_sheet.column_dimensions["B"].width = 76
    ai_sheet.column_dimensions["C"].width = 72

    runtime_sheet = wb.create_sheet("Runtime")
    runtime_sheet.append(["field", "value"])
    style_header(runtime_sheet[1])
    for key, value in runtime_context.items():
        runtime_sheet.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])
    runtime_sheet.column_dimensions["A"].width = 32
    runtime_sheet.column_dimensions["B"].width = 100

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(workbook_path)
    return True


def cell_value(value):
    if isinstance(value, str):
        return value.strip()
    return value


def normalize_bool(value):
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"true", "yes", "y", "1"}:
        return True
    if text in {"false", "no", "n", "0"}:
        return False
    return None


def cleaned_mapping(value):
    return {key: child for key, child in value.items() if child not in (None, "", [], {})}


def validate_text_answer(item_code, value, *, min_words=5):
    text = str(value or "").strip()
    if text.lower() in PLACEHOLDER_TEXT or len(text) < 20 or len(text.split()) < min_words:
        raise ValueError(f"{item_code} needs a more complete written response.")
    return text


def validate_response_value(row):
    item_code = row["item_code"]
    value = row.get("value")
    response_type = row["response_type"]
    if response_type == "likert_5":
        if isinstance(value, str) and value.isdigit():
            value = int(value)
        if not isinstance(value, int) or isinstance(value, bool) or value < 1 or value > 5:
            raise ValueError(f"{item_code} must be an integer from 1 to 5.")
        return value
    if response_type == "yes_no":
        value = normalize_bool(value)
        if not isinstance(value, bool):
            raise ValueError(f"{item_code} must be TRUE or FALSE.")
        return value
    if response_type == "text":
        return validate_text_answer(item_code, value)
    raise ValueError(f"Unsupported response type for {item_code}: {response_type}")


def infer_provider(url):
    host = urlparse(url).netloc.lower().split(":", 1)[0]
    if host.startswith("www."):
        host = host[4:]
    if host == "claude.ai":
        return "claude"
    if host in {"chatgpt.com", "chat.openai.com"}:
        return "chatgpt"
    if host in {"gemini.google.com", "aistudio.google.com"}:
        return "gemini"
    if host == "copilot.microsoft.com":
        return "copilot"
    if host == "perplexity.ai":
        return "perplexity"
    return "other"


def validate_ai_chat_link(chat):
    used_genai = normalize_bool(chat.get("used_genai"))
    if used_genai is not True:
        return None
    chat_url = str(chat.get("chat_url") or "").strip()
    parsed = urlparse(chat_url)
    host = parsed.netloc.lower().split(":", 1)[0]
    if host.startswith("www."):
        host = host[4:]
    if parsed.scheme != "https" or host not in ALLOWED_CHAT_HOSTS:
        raise ValueError("When used_genai is TRUE, AI Chat.chat_url must be an HTTPS shared-chat URL from an allowed provider.")
    if normalize_bool(chat.get("student_confirmed_reviewed")) is not True:
        raise ValueError("Set AI Chat.student_confirmed_reviewed to TRUE after reviewing the shared chat for sensitive information.")
    provider = str(chat.get("provider") or "").strip().lower() or infer_provider(chat_url)
    return {"provider": provider, "chat_url": chat_url, "student_confirmed_reviewed": True}


def sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        records.append({headers[index]: cell_value(value) for index, value in enumerate(row) if index < len(headers)})
    return records


def load_response_workbook(path):
    wb = load_workbook(path, data_only=False)
    survey_rows = sheet_records(wb["Survey"])
    assignment_values = {row["field"]: row.get("value") for row in sheet_records(wb["Assignment"])}
    ai_values = {row["field"]: row.get("value") for row in sheet_records(wb["AI Chat"])}
    runtime_values = {row["field"]: row.get("value") for row in sheet_records(wb["Runtime"])} if "Runtime" in wb.sheetnames else {}
    artifact = {
        "data_pull_summary": {
            "raw_jobs_pulled": assignment_values.get("data_pull_summary.raw_jobs_pulled"),
            "bronze_records_written": assignment_values.get("data_pull_summary.bronze_records_written"),
            "silver_records_written": assignment_values.get("data_pull_summary.silver_records_written"),
            "duplicate_jobs_removed": assignment_values.get("data_pull_summary.duplicate_jobs_removed"),
            "malformed_rows_excluded": assignment_values.get("data_pull_summary.malformed_rows_excluded"),
            "aws_hours_used": assignment_values.get("data_pull_summary.aws_hours_used"),
            "estimated_aws_spend_usd": assignment_values.get("data_pull_summary.estimated_aws_spend_usd"),
        },
        "domain_topic": {
            "target_role": assignment_values.get("domain_topic.target_role"),
            "target_industry": assignment_values.get("domain_topic.target_industry"),
            "target_geography": assignment_values.get("domain_topic.target_geography"),
            "skill_cluster": assignment_values.get("domain_topic.skill_cluster"),
        },
        "selected_resources": [part.strip() for part in str(assignment_values.get("selected_resources") or "").split(",") if part.strip()],
        "assignment_answers": {
            "thought_process": assignment_values.get("assignment_answers.thought_process"),
            "assignment_usefulness": assignment_values.get("assignment_answers.assignment_usefulness"),
            "aws_cost_notes": assignment_values.get("assignment_answers.aws_cost_notes"),
        },
        "output_file_manifest": [
            {"path": part.strip()}
            for part in str(assignment_values.get("output_file_manifest") or "").split(",")
            if part.strip()
        ],
    }
    return {"survey_responses": survey_rows, "ai_chat_link": ai_values, "assignment_artifact": artifact, "runtime": runtime_values}


def validate_assignment_artifact(artifact):
    artifact["data_pull_summary"] = cleaned_mapping(artifact.get("data_pull_summary") or {})
    artifact["domain_topic"] = cleaned_mapping(artifact.get("domain_topic") or {})
    artifact["selected_resources"] = [str(item).strip() for item in artifact.get("selected_resources") or [] if str(item).strip()]
    artifact["output_file_manifest"] = [row for row in artifact.get("output_file_manifest") or [] if row.get("path")]
    answers = cleaned_mapping(artifact.get("assignment_answers") or {})
    if "thought_process" in answers:
        answers["thought_process"] = validate_text_answer("assignment_answers.thought_process", answers["thought_process"])
    if "assignment_usefulness" in answers:
        answers["assignment_usefulness"] = validate_text_answer("assignment_answers.assignment_usefulness", answers["assignment_usefulness"], min_words=3)
    artifact["assignment_answers"] = answers
    return artifact


def validate_workbook(workbook_path, *, assignment_code, notebook_run_id=None, output_json=None):
    workbook = load_response_workbook(workbook_path)
    responses = []
    for row in workbook["survey_responses"]:
        responses.append({"item_code": row["item_code"], "value": validate_response_value(row)})
    ai_chat_link = validate_ai_chat_link(workbook["ai_chat_link"])
    assignment_artifact = validate_assignment_artifact(workbook["assignment_artifact"])
    notebook_run_id = notebook_run_id or str(uuid.uuid4())
    submit_payload = {
        "notebook_run_id": notebook_run_id,
        "notebook_version": NOTEBOOK_VERSION,
        "responses": responses,
        "assignment_artifact": assignment_artifact,
        "client_metadata": {
            "runtime": "research_survey_workbook.py",
            "python_version": platform.python_version(),
            "platform": platform.platform(),
            "runtime_context": workbook.get("runtime", {}),
            "submitted_from_notebook_at": datetime.now(timezone.utc).isoformat(),
        },
    }
    if ai_chat_link:
        submit_payload["ai_chat_link"] = ai_chat_link
    record = {
        "assignment_code": assignment_code,
        "notebook_run_id": notebook_run_id,
        "notebook_version": NOTEBOOK_VERSION,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "submit_to_api": False,
        "server_status_code": None,
        "server_response": None,
        "submitted_payload_without_api_key": submit_payload,
        "response_workbook_path": str(workbook_path),
    }
    output_json = Path(output_json) if output_json else Path(workbook_path).with_name(f"{assignment_code}_submission.json")
    output_json.write_text(json.dumps(record, indent=2), encoding="utf-8")
    return record, output_json


def submit_record(record, assignment_code):
    load_dotenv_file()
    base_url = base_url_from_env()
    prefix = path_prefix_from_env(base_url)
    headers = {
        "X-API-Key": api_key_from_env(),
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    attempts = []
    response = None
    for submit_url in survey_submit_url_candidates(base_url, prefix, assignment_code):
        try:
            response = requests.post(submit_url, headers=headers, json=record["submitted_payload_without_api_key"], timeout=30)
            break
        except requests.exceptions.RequestException as exc:
            attempts.append({"url": submit_url, "error": str(exc)})
    if response is None:
        raise RuntimeError(survey_failure_message(attempts, action="submit to"))
    try:
        body = response.json()
    except ValueError as exc:
        raise RuntimeError(f"Submission failed with status {response.status_code}: {response.text[:500]}") from exc
    record["submit_to_api"] = True
    record["server_status_code"] = response.status_code
    record["server_response"] = body
    return record


def command_init(args):
    survey, survey_url = fetch_survey(args.assignment_code)
    output_dir = Path(args.output_dir)
    workbook_path = output_dir / f"{args.assignment_code}_response_workbook.xlsx"
    created = create_workbook(survey, workbook_path, force=args.force)
    status = "created" if created else "exists"
    print(json.dumps({"status": status, "workbook_path": str(workbook_path), "survey_url": survey_url}, indent=2))


def command_validate(args):
    record, output_json = validate_workbook(
        args.workbook,
        assignment_code=args.assignment_code,
        notebook_run_id=args.notebook_run_id,
        output_json=args.output_json,
    )
    if args.submit:
        record = submit_record(record, args.assignment_code)
        output_json.write_text(json.dumps(record, indent=2), encoding="utf-8")
    print(json.dumps({"validated": True, "output_json": str(output_json), "submitted": bool(args.submit)}, indent=2))


def main():
    parser = argparse.ArgumentParser(description="Create and validate randomized employability research Excel workbooks.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init", help="Fetch randomized questions and create the assignment Excel workbook once.")
    init_parser.add_argument("--assignment-code", required=True)
    init_parser.add_argument("--output-dir", default="research_submissions")
    init_parser.add_argument("--force", action="store_true", help="Overwrite an existing workbook.")
    init_parser.set_defaults(func=command_init)

    validate_parser = subparsers.add_parser("validate", help="Validate a filled Excel workbook and write final JSON.")
    validate_parser.add_argument("--assignment-code", required=True)
    validate_parser.add_argument("--workbook", required=True)
    validate_parser.add_argument("--notebook-run-id", default="")
    validate_parser.add_argument("--output-json", default="")
    validate_parser.add_argument("--submit", action="store_true", help="POST the validated payload to the research API.")
    validate_parser.set_defaults(func=command_validate)

    load_dotenv_file()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
